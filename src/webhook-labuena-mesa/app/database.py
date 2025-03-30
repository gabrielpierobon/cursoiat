import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import random
import string
from dotenv import load_dotenv

load_dotenv()

class DatabaseManager:
    def __init__(self):
        self.db_file = os.getenv("DB_FILE", "data/members.xlsx")
        self.voucher_prefix = os.getenv("VOUCHER_PREFIX", "LBM-")
        self._initialize_db()
    
    def _initialize_db(self):
        # Comprobar si la ruta existe, sino crearla
        os.makedirs(os.path.dirname(self.db_file), exist_ok=True)
        
        # Si el archivo no existe, crearlo con las cabeceras
        if not os.path.exists(self.db_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Miembros"
            ws.append(["Nombre", "Email", "Teléfono", "Fecha Registro", "Voucher", "Redimido"])
            wb.save(self.db_file)
    
    def check_member_exists(self, email, telefono):
        wb = openpyxl.load_workbook(self.db_file)
        ws = wb["Miembros"]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == email or row[2] == telefono:
                return True
        return False
    
    def add_member(self, nombre, email, telefono):
        if self.check_member_exists(email, telefono):
            return {
                "success": False,
                "message": "Este email o teléfono ya está registrado en nuestro club de miembros.",
                "is_new": False
            }
        
        # Generar código voucher
        voucher = self._generate_voucher()
        
        # Añadir miembro
        wb = openpyxl.load_workbook(self.db_file)
        ws = wb["Miembros"]
        
        ws.append([
            nombre,
            email,
            telefono,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            voucher,
            "No"
        ])
        
        wb.save(self.db_file)
        
        return {
            "success": True,
            "message": "¡Registro exitoso en el club de miembros de La Buena Mesa!",
            "voucher_code": voucher,
            "is_new": True
        }
    
    def _generate_voucher(self):
        # Generar código alfanumérico único de 8 caracteres
        chars = string.ascii_uppercase + string.digits
        code = ''.join(random.choice(chars) for _ in range(8))
        return f"{self.voucher_prefix}{code}"
